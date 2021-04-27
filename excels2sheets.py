import scipy.io as sio
import pandas as pd
import os

excel_path='C:/Users/qeebeast7/Documents'
res_file='merge_result'
write_path=excel_path+'/'+res_file+'.xlsx'

for file in os.listdir(excel_path):
    if('.xlsx' in file):
        file_path = os.path.join(excel_path, file)
        df = pd.read_excel(file_path)
        columns = df.columns
        columns = list(df.columns)
        value = [str] * len(columns)
        str_convert = dict(zip(columns, value))
        df = pd.read_excel(file_path,converters=str_convert)
        keys = list(df.keys())
        values = df.values.tolist()
        sheet_name = file.replace('.xlsx', '')
        import openpyxl

        if (os.path.exists(write_path)):
            wb = openpyxl.load_workbook(write_path)
        else:
            wb = openpyxl.Workbook()
        if (sheet_name in wb.sheetnames):
            wb.remove_sheet(wb[sheet_name])
        wb.create_sheet(title=sheet_name, index=0)
        wb[sheet_name].append(keys)
        for value in values:
            wb[sheet_name].append(value)

        for name in wb.sheetnames:
            ws = wb[name]
            if (ws.cell(1, 1).value == None):
                wb.remove_sheet(ws)
        wb.save(write_path)
